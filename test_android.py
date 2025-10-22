import time
import pytest
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.actions import interaction
from selenium.webdriver.common.actions.pointer_input import PointerInput
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException

APP_PACKAGE = "mn.xacbank.teen"
APK_PATH = r"C:\Users\5741\Downloads\xac-android-uat.apk"
EXCEL_PATH = r"C:\Users\5741\Desktop\XacTeenUAT\credentials-uat.xlsx"

qr_icon = chr(0xf433)
icon = chr(0xF002)
icon_char = chr(0xE86D)
back_icon_char = chr(0xF053)
amount = "1000"

@pytest.fixture(scope="module")
def sheet():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.cell(row=1, column=1).value = "Username"
    ws.cell(row=1, column=2).value = "Password"
    ws.cell(row=1, column=3).value = "Bank In Name"
    ws.cell(row=1, column=4).value = "Bank In Account"
    ws.cell(row=1, column=5).value = "Bank Out Name"
    ws.cell(row=1, column=6).value = "Bank Out Account"
    ws.cell(row=1, column=7).value = "PIN"
    ws.cell(row=1, column=8).value = "Nickname"
    ws.cell(row=1, column=9).value = "Account Name"
    ws.cell(row=1, column=10).value = "Phone Number Transfer"
    ws.cell(row=1, column=11).value = "Bank In Transfer Result"
    ws.cell(row=1, column=12).value = "Bank Out Transfer Result"
    ws.cell(row=1, column=13).value = "Phone Number Transfer Result"
    ws.cell(row=1, column=14).value = "QR Transfer Result"
    ws.cell(row=1, column=15).value = "QR Generate Result"
    ws.cell(row=1, column=16).value = "Theme Result"
    ws.cell(row=1, column=17).value = "Nickname Result"
    ws.cell(row=1, column=18).value = "Profile Result"
    ws.cell(row=1, column=19).value = "Device Result"
    ws.cell(row=1, column=20).value = "Login Result"
    yield ws
    wb.save(EXCEL_PATH)
    wb.close()

@pytest.fixture
def platform():
    return "Android"

def navigate_to_home(driver):
    try:
        for _ in range(3):
            driver.back()

        home_btn = driver.find_element(AppiumBy.ID, "mn.xacbank.teen:id/home_button")
        home_btn.click()
    except NoSuchElementException:
        print("Home button not found; used back button to return to Home.")

def is_logged_in(driver):
    home_texts = [
        "Сүүлийн гүйлгээнүүд",
        "Мөнгө илгээх",
        "Үлдэгдэл",
    ]
    for text in home_texts:
        elements = driver.find_elements(
            AppiumBy.ANDROID_UIAUTOMATOR,
            f'new UiSelector().text("{text}")'
        )
        if elements:
            return True
    return False

def logout(driver):
    wait = WebDriverWait(driver, 5)
    try:
        wait.until(EC.element_to_be_clickablsfe(
            (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().className("android.widget.ImageView").instance(3)'))).click()
        wait.until(EC.element_to_be_clickable(
            (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().description("Апп-аас гарах")'))).click()
        time.sleep(1)
    except Exception:
        pass

def ensure_login_screen(driver):
    wait = WebDriverWait(driver, 5)
    try:
        profile_icon = wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().className("android.widget.ImageView").instance(3)')))
        profile_icon.click()
        logout_btn = wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().description("Апп-аас гарах")')))
        logout_btn.click()
        time.sleep(2)
        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().description("Нэвтрэх")')))
        return
    except Exception:
        pass

    try:
        sign_in = wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().description("Нэвтрэх")')))
        sign_in.click()
        time.sleep(1)
    except Exception:
        pass

    try:
        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().text("Сайн уу")')))
        other_user_btn = wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().text("Өөр хэрэглэгч?")')))
        other_user_btn.click()
        time.sleep(1)
    except Exception:
        pass

def qr_t(driver):
    wait = WebDriverWait(driver, 5)
    try:
        qicon = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR, 
            'new UiSelector().description("SCREEN_CAMERA")'
        )))
        qicon.click()

        my_qr = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().description("Миний QR")'
        )))
        my_qr.click()

        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.widget.ImageView").instance(1)'
        )))
        return True
    except Exception as e:
        print(f"QR navigation failed: {e}")
        return False


def enter_pin(driver, pin):
    wait = WebDriverWait(driver, 5)
    pin = str(pin).strip()
    print(f"Entering PIN: {pin}")
    for i, digit in enumerate(pin):
        try:
            key = wait.until(EC.element_to_be_clickable((
                AppiumBy.ANDROID_UIAUTOMATOR,
                f'new UiSelector().text("{digit}").instance(0)'
            )))
            key.click()
            time.sleep(0.15 if i < len(pin) - 1 else 0.5)
        except Exception as e:
            print(f"Error clicking digit '{digit}': {e}")
            raise



def bankin_transaction(driver, acc, pin, amount):
    wait = WebDriverWait(driver, 10)
    pin = str(pin)

    print("[bankin] Clicking Transfer icon...")
    try:
        transfer_icon = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Мөнгө илгээх")'
        )))
        transfer_icon.click()
    except Exception as e:
        print("[bankin_transaction] Error:", e)
        return "Transfer icon not found"

    try:
        print("[bankin] Clicking Bank option...")
        bank_btn = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("ХасБанк")'
        )))
        bank_btn.click()
    except Exception as e:
        print("[bankin_transaction] Bank button error:", e)
        return "Bank button not found"

    try:
        print("[bankin] Inputting Account...")
        acc_input = wait.until(EC.presence_of_element_located((
            AppiumBy.CLASS_NAME, "android.widget.EditText"
        )))
        acc_input.send_keys(acc)

        srch_btn = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().description("BUTTON_SEARCH_FOR_ACCOUNT")'
        )))
        srch_btn.click()

        icon_btn = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().className("android.widget.FrameLayout").instance(5)'
        )))
        icon_btn.click()

        print("[bankin] Inputting Amount via digit buttons...")
        for digit in str(amount):
            wait.until(EC.element_to_be_clickable((
                AppiumBy.ANDROID_UIAUTOMATOR, 
                f'new UiSelector().description("AMOUNT_{digit}")'
            ))).click()
            time.sleep(0.2)

        send = driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Илгээх")')
        send.click()

        print("[bankin] Entering PIN...")
        for digit in str(pin):
            wait.until(EC.element_to_be_clickable((
                AppiumBy.ANDROID_UIAUTOMATOR,
                f'new UiSelector().description("PIN_{digit}")'
            ))).click()
            time.sleep(0.2)


        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Амжилттай")'
        )))
        print("[bankin] Transfer successful!")

        ok_btn = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().text("Окэй")')))
        ok_btn.click()

        return "Success"

    except Exception as e:
        print("[bankin_transaction] Error during transaction:", e)
        return "Failed"




def bankout_transaction(driver, bank, account, pin, ic, amount):
    wait = WebDriverWait(driver, 5)
    result = False
    try:
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Мөнгө илгээх")'
        ))).click()

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Банк хооронд")'
        ))).click()

        account_num = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("IBAN дугаар")'
        )))
        account_num.clear()
        account_num.send_keys(account)

        search_btn = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR, f'new UiSelector().text("{ic}")')))
        search_btn.click()

        account_name = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Дансны нэр гараас оруулах")'
        )))
        account_name.clear()
        account_name.send_keys(bank)

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Үргэлжүүлэх")'
        ))).click()

        print("[bankout] Inputting Amount via digit buttons...")
        for digit in str(amount):
            wait.until(EC.element_to_be_clickable((
                AppiumBy.ANDROID_UIAUTOMATOR, 
                f'new UiSelector().description("AMOUNT_{digit}")'
            ))).click()
            time.sleep(0.2)

        send = driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Илгээх")')
        send.click()

        print("[bankout] Entering PIN...")
        for digit in str(pin):
            wait.until(EC.element_to_be_clickable((
                AppiumBy.ANDROID_UIAUTOMATOR,
                f'new UiSelector().description("PIN_{digit}")'
            ))).click()
            time.sleep(0.2)

        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().text("Амжилттай")')))
        print("Transfer successful!")

        ok_btn = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().text("Окэй")')))
        ok_btn.click()

        result = True
    except Exception:
        result = False
    return result

def qr_transaction(driver, pin, amount):
    wait = WebDriverWait(driver, 10)
    try:
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().description("SCREEN_CAMERA")'
        ))).click()

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().description("BUTTON_IMAGE_UPLOAD")'
        ))).click()

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().resourceId("com.google.android.providers.media.module:id/icon_thumbnail").instance(2)'
        ))).click()

        print("[qr] Inputting Amount via digit buttons...")
        for digit in str(amount):
            wait.until(EC.element_to_be_clickable((
                AppiumBy.ANDROID_UIAUTOMATOR, 
                f'new UiSelector().description("AMOUNT_{digit}")'
            ))).click()
            time.sleep(0.2)

        send = driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Төлөх")')
        send.click()

        print("[qr] Entering PIN...")
        for digit in str(pin):
            wait.until(EC.element_to_be_clickable((
                AppiumBy.ANDROID_UIAUTOMATOR,
                f'new UiSelector().description("PIN_{digit}")'
            ))).click()
            time.sleep(0.2)

        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().text("Амжилттай")')))
        print("Transfer successful!")

        ok_btn = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().text("Окэй")')))
        ok_btn.click()

        result = True
    except Exception:
        result = False
    return result

def pn_transaction(driver, pnum, pin, ic, amount):
    wait = WebDriverWait(driver, 5)
    result = False
    try:
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Мөнгө илгээх")'
        ))).click()

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Утасны дугаар")'
        ))).click()

        edit_field = wait.until(EC.element_to_be_clickable((
            AppiumBy.CLASS_NAME,
            "android.widget.EditText"
        )))
        edit_field.clear()
        edit_field.send_keys(pnum)

        search_btn = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR, f'new UiSelector().text("{ic}")')))
        search_btn.click()

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.view.View").instance(3)'
        ))).click()

        print("[pn] Inputting Amount via digit buttons...")
        for digit in str(amount):
            wait.until(EC.element_to_be_clickable((
                AppiumBy.ANDROID_UIAUTOMATOR, 
                f'new UiSelector().description("AMOUNT_{digit}")'
            ))).click()
            time.sleep(0.2)

        send = driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Илгээх")')
        send.click()

        print("[pn] Entering PIN...")
        for digit in str(pin):
            wait.until(EC.element_to_be_clickable((
                AppiumBy.ANDROID_UIAUTOMATOR,
                f'new UiSelector().description("PIN_{digit}")'
            ))).click()
            time.sleep(0.2)

        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().text("Амжилттай")')))
        print("Transfer successful!")

        ok_btn = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().text("Окэй")')))
        ok_btn.click()

        result = True
    except Exception:
        result = False
    return result


def card_t(driver, pin):
    try:
        wait = WebDriverWait(driver, 5)
        print("Starting card test...")

        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.widget.ImageView").instance(4)'
        ))).click()
        print("Clicked ImageView.instance(4)")

        time.sleep(1)

        try:
            wait.until(EC.presence_of_element_located((
                AppiumBy.ANDROID_UIAUTOMATOR,
                'new UiSelector().className("android.widget.ImageView").instance(0)'
            ))).click()
            print("Clicked ImageView.instance(0)")
        except Exception as e:
            print(f"ImageView.instance(0) not found or not clickable, trying generic ImageView... ({e})")
            wait.until(EC.presence_of_element_located((
                AppiumBy.ANDROID_UIAUTOMATOR,
                'new UiSelector().className("android.widget.ImageView")'
            ))).click()
            print("Clicked generic ImageView")

        wait.until(EC.element_to_be_clickable((
            AppiumBy.XPATH,
            '//android.widget.TextView[@text="Мэдээлэл харах"]'
        ))).click()

        print('Clicked "Мэдээлэл харах"')


        print("[card] Entering PIN...")
        for digit in str(pin):
            pin_btn = wait.until(EC.element_to_be_clickable((
                AppiumBy.XPATH,
                f'//android.widget.TextView[@text="{digit}"]'
            )))
            pin_btn.click()
            time.sleep(0.2)

        print("Entered PIN")

        wait.until(EC.element_to_be_clickable((
            AppiumBy.XPATH,
            '//android.widget.TextView[@text="Мэдээлэл нуух"]'
        ))).click()

        print('Clicked "Мэдээлэл нуух"')

        wait.until(EC.element_to_be_clickable((
            AppiumBy.XPATH,
            '//android.widget.TextView[@text="Картын пин авах"]'
        ))).click()

        print('Clicked "Картын пин авах"')


        wait.until(EC.element_to_be_clickable((
            AppiumBy.XPATH,
            '//android.widget.TextView[@text="Тийм"]'
        ))).click()

        print('Clicked "Тийм"')

        print("[card] Entering PIN...")
        for digit in str(pin):
            pin_btn = wait.until(EC.element_to_be_clickable((
                AppiumBy.XPATH,
                f'//android.widget.TextView[@text="{digit}"]'
            )))
            pin_btn.click()
            time.sleep(0.2)

        print("Entered PIN")

        wait.until(EC.element_to_be_clickable((
            AppiumBy.XPATH,
            '//android.widget.TextView[@text="Окей"]'
        ))).click()

        print('Clicked "Окей"')

        wait.until(EC.element_to_be_clickable((
            AppiumBy.XPATH,
            '//android.widget.TextView[@text="Картыг түр хаах"]'
        ))).click()

        print('Clicked "Картыг түр хаах"')

        print("[card] Entering PIN...")
        for digit in str(pin):
            pin_btn = wait.until(EC.element_to_be_clickable((
                AppiumBy.XPATH,
                f'//android.widget.TextView[@text="{digit}"]'
            )))
            pin_btn.click()
            time.sleep(0.2)

        print("Entered PIN")

        time.sleep(10)

        wait.until(EC.element_to_be_clickable((
            AppiumBy.XPATH,
            '//android.widget.TextView[@text="Картыг нээх"]'
        ))).click()

        print('Clicked "Картыг нээх"')

        print("[card] Entering PIN...")
        for digit in str(pin):
            pin_btn = wait.until(EC.element_to_be_clickable((
                AppiumBy.XPATH,
                f'//android.widget.TextView[@text="{digit}"]'
            )))
            pin_btn.click()
            time.sleep(0.2)

        print("Entered PIN")

        time.sleep(10)

        print("card test completed successfully.")
        return True
    except Exception:
        return False

def go_to_settings(driver):
    wait = WebDriverWait(driver, 5)
    try:
        elems = wait.until(EC.presence_of_all_elements_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Миний профайл")')))
        if elems:
            return True
    except Exception:
        pass
    try:
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.widget.ImageView").instance(3)'))).click()
        wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Миний профайл")')))
        return True
    except Exception:
        return False


def set_nickname(driver, new_nickname):
    wait = WebDriverWait(driver, 5)

    try:
        el4 = driver.find_element(AppiumBy.ACCESSIBILITY_ID, "SCREEN_NICKNAME_UPDATE")
        el4.click()

        nickname_field = wait.until(EC.element_to_be_clickable(
            (AppiumBy.CLASS_NAME, "android.widget.EditText")))
        nickname_field.clear()
        nickname_field.send_keys(new_nickname)

        save_btn = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Хадгалах")')))
        save_btn.click()

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().description("BUTTON_NAVIGATION_BACK")'
        ))).click()

        return True

    except Exception as e:
        print(f"[set_nickname] Error: {e}")
        return False


def change_profile(driver):
    wait = WebDriverWait(driver, 5)
    try:
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.widget.ImageView").instance(1)'
        ))).click()
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().resourceId("com.google.android.providers.media.module:id/icon_thumbnail").instance(1)'
        ))).click()
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().resourceId("mn.xacbank.teen:id/crop_image_menu_crop")'
        ))).click()
        return True
    except Exception:
        return False

def swipe_left(driver):
    size = driver.get_window_size()
    y = int(size['height'] * 0.92)
    start_x = int(size['width'] * 0.8)
    end_x = int(size['width'] * 0.2)
    driver.swipe(start_x, y, end_x, y, 500)
    time.sleep(0.5)

def open_theme_picker(driver):
    wait = WebDriverWait(driver, 7)
    try:
        if driver.find_elements(
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Theme солих")'
        ):
            return True
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.widget.ImageView").instance(3)'
        ))).click()
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiScrollable(new UiSelector().scrollable(true)).scrollIntoView(new UiSelector().text("Theme солих"))'
        ))).click()
        return True
    except Exception:
        return False

def long_press(driver, element, duration=2000):
    """Reusable long-press helper (duration in ms)"""
    actions = ActionChains(driver)
    finger = PointerInput(interaction.POINTER_TOUCH, "finger")

    builder = ActionBuilder(driver, mouse=finger)
    builder.pointer_action.move_to(element)
    builder.pointer_action.pointer_down()
    builder.pointer_action.pause(duration / 1000)
    builder.pointer_action.pointer_up()
    builder.perform()

def friends_circle(driver):
    wait = WebDriverWait(driver, 5)
    try:
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().description("TAB_MY_CIRCLE")'
        ))).click()
        time.sleep(0.5)

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.view.ViewGroup").instance(49)'
        ))).click()
        time.sleep(0.5)

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Нэмэх")'
        ))).click()
        time.sleep(0.5)

        img_element = wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.widget.ImageView").instance(2)'
        )))
        time.sleep(0.5)

        long_press(driver, img_element, 2000)

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Хүрээллээс хасах")'
        ))).click()

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Хасах")'
        ))).click()

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().description("TAB_HOME")'
        ))).click()

        return True
    except Exception as e:
        print(f"Failed to open Friends Circle: {e}")
        return False


def theme_t(driver):
    wait = WebDriverWait(driver, 20)
    state = False

    # Step 1: Open the theme picker
    if not open_theme_picker(driver):
        print("Failed to open theme picker.")
        return state
    print("Opened theme picker.")

    # Step 2: Click "Зураг оруулах"
    try:
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Зураг оруулах")'
        ))).click()
        print("Clicked 'Зураг оруулах'")
    except Exception as e:
        print(f"Failed to click 'Зураг оруулах': {e}")
        return state

    # Step 3: Click thumbnail image
    try:
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().resourceId("com.google.android.providers.media.module:id/icon_thumbnail").instance(0)'
        ))).click()
        print("Clicked image thumbnail")
    except Exception as e:
        print(f"Failed to click image thumbnail: {e}")
        return state

    # Step 4: Click crop button
    try:
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().resourceId("mn.xacbank.teen:id/crop_image_menu_crop")'
        ))).click()
        print("Clicked crop button")
    except Exception as e:
        print(f"Failed to click crop button: {e}")
        return state

    time.sleep(15)

    # Step 5: Scroll to and click "Theme солих"
    try:
        element = wait.until(EC.presence_of_element_located((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiScrollable(new UiSelector().scrollable(true)).scrollIntoView(new UiSelector().text("Theme солих"))'
        )))
        element.click()
        print("Clicked 'Theme солих'")
    except Exception as e:
        print(f"Failed to scroll/click 'Theme солих': {e}")
        return state


    # Step 6: Select built-in theme and confirm
    try:
        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().className("android.widget.ImageView").instance(6)'
        ))).click()

        print("Clicked built-in theme")

        wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR,
            'new UiSelector().text("Сонгох")'
        ))).click()
        print("Clicked 'Сонгох'")

        state = True
    except Exception as e:
        print(f"Failed to select and confirm theme: {e}")
        state = False

    return state



def test_main(driver, sheet, platform):
    wait = WebDriverWait(driver, 3)
    driver.activate_app(APP_PACKAGE)
    row = 3

    while True:
        ensure_login_screen(driver)

        user = sheet.cell(row=row, column=1).value
        pwd = sheet.cell(row=row, column=2).value
        bank = sheet.cell(row=row, column=3).value
        acc = sheet.cell(row=row, column=4).value
        bankout_name = sheet.cell(row=row, column=5).value
        bankout_acc = sheet.cell(row=row, column=6).value
        pin = str(sheet.cell(row=row, column=7).value)
        nickname = sheet.cell(row=row, column=8).value
        bankout_usr_name = sheet.cell(row=row, column=9).value
        ph = sheet.cell(row=row, column=10).value

        if not user or not pwd:
            break

        try:
            username_field = wait.until(EC.presence_of_element_located((
                AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().className("android.widget.EditText").instance(0)')))
            username_field.clear()
            username_field.send_keys(user)
        except Exception:
            sheet.cell(row=row, column=20).value = "login failed"
            row += 1
            continue

        try:
            password_field = wait.until(EC.presence_of_element_located((
                AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().className("android.widget.EditText").instance(1)')))
            password_field.clear()
            password_field.send_keys(pwd)
        except Exception:
            sheet.cell(row=row, column=20).value = "login failed"
            row += 1
            continue

        try:
            login_btn = wait.until(EC.element_to_be_clickable((
                AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().description("Нэвтрэх")')))
            login_btn.click()
        except Exception:
            print("Login button click failed.")

        time.sleep(2)

        if not is_logged_in(driver):
            sheet.cell(row=row, column=20).value = "login failed"
            row += 1
            driver.terminate_app(APP_PACKAGE)
            time.sleep(0.3)
            driver.activate_app(APP_PACKAGE)
            time.sleep(0.3)
            continue
        else:
            sheet.cell(row=row, column=20).value = "login passed"

        # # BankIn Transfer Test
        # transfer_result = bankin_transaction(driver, acc, pin, amount)
        # if transfer_result:
        #    sheet.cell(row=row, column=11).value = "BankIn transfer passed"
        # else:
        #    sheet.cell(row=row, column=11).value = "BankIn transfer failed"
        # time.sleep(0.3)

        # # BankOut Transfer Test
        # transfer_result = bankout_transaction(driver, bankout_usr_name, bankout_acc, pin, icon, amount)
        # if transfer_result:
        #    sheet.cell(row=row, column=12).value = "BankOut transfer passed"
        # else:
        #    sheet.cell(row=row, column=12).value = "BankOut transfer failed"
        # time.sleep(0.3)

        # Phone Number Transfer Test
        transfer_result = pn_transaction(driver, ph, pin, icon, amount)
        if transfer_result:
           sheet.cell(row=row, column=13).value = "Phone number transfer passed"
        else:
           sheet.cell(row=row, column=13).value = "Phone number transfer failed"
        time.sleep(0.3)

        # QR Transfer Test
        transfer_result = qr_transaction(driver, pin, amount)
        if transfer_result:
           sheet.cell(row=row, column=14).value = "QR transfer passed"
        else:
           sheet.cell(row=row, column=14).value = "QR transfer failed"
        time.sleep(0.3)

        #Card Test
        card_result = card_t(driver, pin)
        if card_result:
            sheet.cell(row=row, column=21).value = "Card passed"
        else:
            sheet.cell(row=row, column=21).value = "Card failed"

        back_btn = wait.until(EC.presence_of_element_located((
            AppiumBy.XPATH, '//android.view.ViewGroup[@content-desc="BUTTON_NAVIGATION_BACK"]'
        )))
        back_btn.click()


        print('Clicked "Back"')

        time.sleep(5)

        back_btn = wait.until(EC.presence_of_element_located((
            AppiumBy.XPATH, '//android.view.ViewGroup[@content-desc="BUTTON_NAVIGATION_BACK"]'
        )))
        back_btn.click()

        print('Clicked "Back"')

        # Friend Test
        if friends_circle(driver):
            sheet.cell(row=row, column=22).value = "Friends circle passed"
        else:
            sheet.cell(row=row, column=22).value = "Friends circle failed"

        # QR Test
        if qr_t(driver):
            sheet.cell(row=row, column=15).value = "QR passed"
        else:
            sheet.cell(row=row, column=15).value = "QR failed"

        qr_back = wait.until(EC.element_to_be_clickable((
            AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().description("BUTTON_NAVIGATION_BACK")'
        )))
        qr_back.click()

        # Theme Test
        if theme_t(driver):
            sheet.cell(row=row, column=16).value = "Theme passed"
        else:
            sheet.cell(row=row, column=16).value = "Theme failed"
        time.sleep(1)

        # Nickname Test
        go_to_settings(driver)
        if set_nickname(driver, nickname):
            sheet.cell(row=row, column=17).value = "Nickname passed"
        else:
            sheet.cell(row=row, column=17).value = "Nickname failed"
        time.sleep(1)

        # Change Profile Test
        go_to_settings(driver)
        if change_profile(driver):
            sheet.cell(row=row, column=18).value = "Profile passed"
        else:
            sheet.cell(row=row, column=18).value = "Profile failed"

        # Platform-specific result
        if platform == "Android":
            sheet.cell(row=row, column=19).value = "Android passed"
        else:
            sheet.cell(row=row, column=19).value = "IOS passed"

        break

        # Prepare for next iteration
        logout(driver)
        driver.terminate_app(APP_PACKAGE)
        time.sleep(1)
        driver.activate_app(APP_PACKAGE)
        time.sleep(1)
        row += 1
